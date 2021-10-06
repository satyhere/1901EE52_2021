import csv
import os
from openpyxl import Workbook

grade_dict = {'AA': 10, 'AB': 9, 'BB': 8, 'BC': 7, 'CC': 6, 'CD': 5, 'DD': 4, 'F': 0, 'I': 0}
i = 0

nameRoll = {}
branch = {}
roll_subjects = {}

subject = {}
subject_name = {}
subject_ltp = {}
subject_credit = {}
with open('names-roll.csv', newline='') as data:
    reader = csv.DictReader(data)
    for i in reader:
        nameRoll[i['Roll']] = i['Name']
        branch[i['Roll']] = i['Roll'][4:6]

with open('subjects_master.csv', newline='') as data:
    reader = csv.DictReader(data)
    for i in reader:
        subject[i['subno']] = [i['subname'], i['ltp'], i['crd']]

with open('grades.csv', newline='') as data:
    reader = csv.DictReader(data)
    for i in reader:
        roll = i['Roll']
        semester = int(i['Sem'])
        subjectCode = i['SubCode']
        subjectType = i['Sub_Type']
        credit = int(i['Credit'])
        grade = i['Grade']
        if roll not in roll_subjects.keys():
            roll_subjects[roll] = {}
            roll_subjects[roll][semester] = []
        elif semester not in roll_subjects[roll].keys():
            roll_subjects[roll][semester] = []

        roll_subjects[roll][semester].append([credit, grade, subjectCode, subjectType])

Result = {}
semesterWise = {}

for roll, s in roll_subjects.items():
    for semester, info in s.items():

        credits = []
        grade = []
        subjectCode = []
        subjectType = []
        for i in info:
            credits.append(i[0])
            grade.append(i[1])
            subjectCode.append(i[2])
            subjectType.append(i[3])

        total = sum(credits)
        p = 0
        for i in range(len(grade)):
            grade[i] = grade[i].strip()

            if grade[i].endswith('*'):
                grade[i] = grade[i][:-1]

            p += grade_dict[grade[i]] * credits[i]

        spi = p / total
        spi = round(spi, 2)
        if roll not in Result.keys():
            Result[roll] = []
            semesterWise[roll] = []

        semesterWise[roll].append([semester, subjectCode, subjectType, grade])
        Result[roll].append([semester, total, spi])

os.mkdir('output')
for r_no, s in Result.items():
    wb = Workbook()
    path = 'output/' + r_no + ".xlsx"
    wb1 = wb.create_sheet('Overall', 0)
    wb1.append(["Roll No", r_no])
    wb1.append(["Name of Student", nameRoll[r_no]])
    wb1.append(["Discipline", branch[r_no]])
    wb1.append(["Semester No.", 1, 2, 3, 4, 5, 6, 7, 8])

    SPI = []
    semesterWiseCredits = []
    for i in s:
        SPI.append(i[2])
        semesterWiseCredits.append(i[1])

    CPI = []
    semesterWiseTotal = []
    t = 0
    c = 0
    for i in range(len(semesterWiseCredits)):
        t += semesterWiseCredits[i]
        c += semesterWiseCredits[i] * SPI[i]

        temp = c / t
        temp = round(temp, 2)
        CPI.append(temp)
        semesterWiseTotal.append(t)

    columns = ["Semester-Wise Credit Taken", "SPI", "Total Credits Taken", "CPI"]
    tags = {0: SPI, 1: semesterWiseCredits, 2: semesterWiseTotal, 3: CPI}

    for i, name in enumerate(columns):
        a = [name] + tags[i]
        wb1.append(a)

    for i in range(len(semesterWise[r_no])):

        wb2 = wb.create_sheet('SEM ' + str(i + 1), i + 1)

        fixed_line = ["Sl No.", "Subject No.", "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade"]
        wb2.append(fixed_line)
        kl = semesterWise[r_no][i]
        for j in range(len(semesterWise[r_no][i][2])):
            wb2.append(
                [j + 1, kl[1][j], subject[kl[1][j]][0], subject[kl[1][j]][1], subject[kl[1][j]][2], kl[2][j], kl[3][j]])

    wb.save(path)
